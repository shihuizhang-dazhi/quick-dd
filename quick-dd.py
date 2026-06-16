#!/usr/bin/env python3
"""
Quick-DD 资产发现工具
用法: python3 run.py <domain>
输出: 一个 Excel 文件，包含如下 sheet：
    1. assets      - 探活后的资产结果
    2. ip_domains  - rdnsdb 的 IP 对应域名
    3. asn_info    - ASN / CIDR / 网段信息
    4. fofa        - FOFA 资产查询结果（需配置 FOFA API）
    5. port_scan   - 存活主机端口扫描结果
"""

import io
import os
import platform
import re
import socket
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import redirect_stdout

from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from domain import fetch_subdomains, probe_host
from real_ip_finder import (
    get_cidr_from_ip138,
    get_domains_from_rdnsdb,
    get_ips_from_ip138,
    get_ip_domain_rows_from_rdnsdb,
    # get_live_ips_from_chapangzhan removed per request
    get_prefixes_for_domain,
    is_private,
)
from subdomain_brute import (
    load_dict, resolve as brute_resolve,
    load_dns_servers, wildcard_test, batch_dns_resolve, load_next_sub,
)
import asyncio

try:
    from ct_lookup import get_subdomains_from_crt
except Exception:
    def get_subdomains_from_crt(domain):
        return []

try:
    from fofa import search_domain as fofa_search_domain
except Exception:
    def fofa_search_domain(domain, size=100):
        return []

try:
    from port_scan import scan_hosts as port_scan_hosts, scan_host as port_scan_host, COMMON_PORTS
except Exception:
    COMMON_PORTS = [80, 443, 22, 21, 25, 3306, 8080, 8443, 8888, 9000]
    def port_scan_host(host, mode="tcp", ports=None):
        return host, [], {}
    def port_scan_hosts(hosts, mode="tcp", ports=None):
        return {h: (h, [], {}) for h in hosts}


# Windows ANSI 支持
if platform.system() == "Windows":
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)
    except Exception:
        pass

# 默认启用的扩展功能
USE_CT = True
USE_ASN = True
USE_FOFA = False
USE_PORT_SCAN = False

C = "\033[96m"
G = "\033[92m"
Y = "\033[93m"
R = "\033[0m"
BOLD = "\033[1m"
RED = "\033[91m"
DIM = "\033[2m"

BANNER = rf"""
{C} 
                     __            __                     __        __ 
                    /  |          /  |                   /  |      /  |
  ______   __    __ $$/   _______ $$ |   __          ____$$ |  ____$$ |
 /      \ /  |  /  |/  | /       |$$ |  /  |______  /    $$ | /    $$ |
/$$$$$$  |$$ |  $$ |$$ |/$$$$$$$/ $$ |_/$$//      |/$$$$$$$ |/$$$$$$$ |
$$ |  $$ |$$ |  $$ |$$ |$$ |      $$   $$< $$$$$$/ $$ |  $$ |$$ |  $$ |
$$ \__$$ |$$ \__$$ |$$ |$$ \_____ $$$$$$  \        $$ \__$$ |$$ \__$$ |
$$    $$ |$$    $$/ $$ |$$       |$$ | $$  |       $$    $$ |$$    $$ |
 $$$$$$$ | $$$$$$/  $$/  $$$$$$$/ $$/   $$/         $$$$$$$/  $$$$$$$/ 
      $$ |                                                             
      $$ |                                                             
      $$/                                       {R}
{DIM}  Quick-DD · 资产发现 · Subdomain Recon{R}
"""


def silent(func, *args, **kwargs):
    """静默执行函数，屏蔽内部 print"""
    with redirect_stdout(io.StringIO()):
        return func(*args, **kwargs)


def _step(label: str):
    """打印步骤标题，左对齐固定宽度"""
    print(f"  {DIM}▸{R} {BOLD}{label}{R}", end="", flush=True)


def _ok(msg: str):
    """打印成功结果（右对齐到步骤行）"""
    print(f" {G}{msg}{R}", flush=True)


def _warn(msg: str):
    """打印警告/跳过"""
    print(f" {Y}{msg}{R}", flush=True)


def _fail(msg: str):
    """打印失败"""
    print(f" {RED}{msg}{R}", flush=True)


def _info(msg: str):
    """打印信息"""
    print(f" {msg}", flush=True)


def _public_ipv4s(domain: str):
    ips = []
    try:
        infos = socket.getaddrinfo(domain, None, socket.AF_INET)
        for info in infos:
            ip = info[4][0]
            if is_private(ip):
                continue
            ips.append(ip)
    except Exception:
        pass
    return list(dict.fromkeys(ips))


def _parse_ports(ports_str: str) -> list:
    """解析端口列表，支持范围如 1-1024 和逗号分隔如 80,443,8080"""
    ports = []
    for part in ports_str.split(","):
        part = part.strip()
        if "-" in part:
            try:
                start, end = part.split("-", 1)
                start, end = int(start.strip()), int(end.strip())
                if 1 <= start <= end <= 65535:
                    ports.extend(range(start, end + 1))
            except ValueError:
                pass
        else:
            try:
                p = int(part)
                if 1 <= p <= 65535:
                    ports.append(p)
            except ValueError:
                pass
    return sorted(set(ports)) if ports else None


def _preferred_url_for_host(host: str, url_candidates):
    if f"https://{host}" in url_candidates:
        return f"https://{host}"
    if f"http://{host}" in url_candidates:
        return f"http://{host}"
    return sorted(url_candidates)[0] if url_candidates else ""


def _expand_cidr(cidr_str: str) -> list:
    """将 CIDR 网段展开为 IP 列表，限制最多 65536 个 IP"""
    import ipaddress
    try:
        network = ipaddress.ip_network(cidr_str, strict=False)
        if network.num_addresses > 65536:
            print(f"  [!] 网段过大 ({network.num_addresses} 个 IP)，限制为 /16 (65536)")
            network = ipaddress.ip_network(f"{network.network_address}/16", strict=False)
        return [str(ip) for ip in network.hosts()]
    except Exception as e:
        print(f"  [!] CIDR 解析失败: {e}")
        return []


def main():
    import argparse
    parser = argparse.ArgumentParser(
        prog="quick-dd.py",
        description=(
            "Quick-DD 资产发现工具\n"
            "输入域名，自动完成：子域采集 → 存活探测 → 端口扫描 → 输出报告"
        ),
        epilog=(
            "示例:\n"
            "  python quick-dd.py example.com                          # 基础扫描\n"
            "  python quick-dd.py example.com --fofa --port-scan       # 全开\n"
            "  python quick-dd.py example.com --port-scan --ports 1-1024  # 自定义端口\n"
            "  python quick-dd.py example.com --fofa-query 'title=\"登录\"'\n"
            "\n"
            "默认功能:\n"
            "  子域采集 (ip138 + 字典爆破 + CT日志)\n"
            "  并发探活 (HTTP/HTTPS)\n"
            "  IP/ASN 查询\n"
            "\n"
            "可选功能:\n"
            "  --fofa       FOFA 资产查询 (需配置 config.ini [fofa])\n"
            "  --port-scan  TCP 端口扫描 (默认 Top100，可用 --ports 自定义)"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        add_help=True,
    )
    parser.add_argument("domain", help="目标域名（如 example.com）")
    parser.add_argument("--fofa", action="store_true", help="启用 FOFA 资产查询")
    parser.add_argument("--port-scan", action="store_true", help="启用 TCP 端口扫描")
    parser.add_argument("--cidr-scan", default="", metavar="CIDR", help="直接扫描 CIDR 网段（如 211.64.160.0/19）")
    parser.add_argument("--fofa-query", default="", metavar="QUERY", help="自定义 FOFA 查询语句（如 'title=\"登录\"'）")
    parser.add_argument("--fofa-size", type=int, default=100, metavar="N", help="FOFA 查询条数（默认 100）")
    parser.add_argument("--ports", default="", metavar="PORTS", help="自定义端口（如 1-1024 或 80,443,8080）")
    args = parser.parse_args()

    global USE_FOFA, USE_PORT_SCAN
    if args.fofa:
        USE_FOFA = True
    if args.port_scan:
        USE_PORT_SCAN = True
    if args.cidr_scan:
        USE_PORT_SCAN = True

    domain = args.domain.strip()
    domain = re.sub(r'^https?://', '', domain).rstrip('/')
    base_domain = re.sub(r'^www\.', '', domain)
    # 支持 PyInstaller 单文件模式：打包资源会被解到临时目录（_MEIPASS），
    # 但输出应写到用户当前工作目录而非临时目录。
    if getattr(sys, "frozen", False):
        bundle_dir = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    else:
        bundle_dir = os.path.dirname(os.path.abspath(__file__))
    # 输出到当前工作目录（执行 exe 的位置或用户指定的工作路径）
    script_dir = os.getcwd()
    output_dir = os.path.join(script_dir, f"{base_domain}_output")
    os.makedirs(output_dir, exist_ok=True)
    xlsx_file = os.path.join(output_dir, f"{base_domain}_assets.xlsx")

    print(BANNER)

    # ── 1. 子域来源采集 ──
    _step("子域采集 (ip138)")
    api_subs = silent(fetch_subdomains, base_domain)
    _ok(f"{len(api_subs)} 个子域名")

    # ── 2. 字典爆破（异步DNS + 泛解析检测 + 递归）──
    dns_servers = load_dns_servers()

    # 泛解析检测
    _step("泛解析检测")
    wildcard_ips = set()
    wildcard_hash = None
    try:
        is_wc, wc_ips, wc_hash = asyncio.run(wildcard_test(base_domain, dns_servers))
        if is_wc:
            wildcard_ips = wc_ips
            wildcard_hash = wc_hash
            _ok(f"泛解析 → {', '.join(sorted(wc_ips)[:3])}...")
        else:
            _ok("未检测到")
    except Exception:
        _warn("检测失败，跳过")

    _step("字典爆破")
    words = load_dict(files=["subnames.txt", "subnames_full.txt"])

    def _brute_progress(done, total):
        msg = f"  进度: {done}/{total}"
        print(f"\r{msg:<50}", end="", flush=True)

    # 异步批量 DNS 解析（500 并发，泛解析过滤 + HTTP二次验证）
    brute_resolved = asyncio.run(
        batch_dns_resolve(
            [f"{w}.{base_domain}" for w in words],
            dns_servers,
            concurrency=500,
            wildcard_ips=wildcard_ips or None,
            wildcard_hash=wildcard_hash or None,
            progress_callback=_brute_progress,
        )
    )
    brute_hosts = set(brute_resolved.keys())

    # 下一级递归爆破（限制父域名数量，避免任务爆炸）
    next_subs = load_next_sub(full=True)
    max_next_parents = 500  # 最多对 500 个父域名做递归
    if next_subs and brute_hosts:
        top_parents = sorted(brute_hosts)[:max_next_parents]
        next_tasks = []
        for parent in top_parents:
            for sub in next_subs:
                next_tasks.append(f"{sub}.{parent}")
        next_tasks = [h for h in next_tasks if h not in brute_resolved]
        print(f"\r{'':<50}", end="", flush=True)  # 清除进度行
        _step(f"二层递归 ({len(top_parents)} 父域名)")
        if next_tasks:
            next_resolved = asyncio.run(
                batch_dns_resolve(
                    next_tasks,
                    dns_servers,
                    concurrency=200,
                    wildcard_ips=wildcard_ips or None,
                    wildcard_hash=wildcard_hash or None,
                    progress_callback=_brute_progress,
                    timeout=2,
                )
            )
            print(f"\r{'':<50}", end="", flush=True)
            _ok(f"+{len(next_resolved)} 新增")
            brute_resolved.update(next_resolved)
        else:
            _ok("无新任务")

    api_set = set(api_subs)
    brute_only = set(brute_resolved.keys()) - api_set
    all_subs = api_set | set(brute_resolved.keys())
    _ok(f"+{len(brute_only)} 新增 ({len(all_subs)} 总计)")

    # ── 3. 证书透明度补全 ──
    if USE_CT:
        _step("CT 日志 (crt.sh)")
        try:
            ct_subs = silent(get_subdomains_from_crt, base_domain)
            ct_only = set(ct_subs) - all_subs
            if ct_subs:
                all_subs |= set(ct_subs)
                _ok(f"+{len(ct_only)} 新增 ({len(all_subs)} 总计)")
            else:
                _warn("无结果")
        except Exception:
            _fail("查询失败")

    # ── 4. FOFA 资产查询 ──
    fofa_results = []
    if USE_FOFA:
        _step("FOFA 资产查询")
        try:
            if args.fofa_query:
                from fofa import search as fofa_search_custom
                fofa_results = silent(fofa_search_custom, args.fofa_query, args.fofa_size)
            else:
                fofa_results = silent(fofa_search_domain, base_domain, args.fofa_size)
            if fofa_results:
                fofa_hosts = set()
                for r in fofa_results:
                    h = r.get("host") or r.get("domain") or ""
                    h = h.strip().lstrip("*.")
                    if h and (h.endswith("." + base_domain) or h == base_domain):
                        fofa_hosts.add(h)
                fofa_only = fofa_hosts - all_subs
                all_subs |= fofa_hosts
                _ok(f"+{len(fofa_only)} 新增 ({len(fofa_results)} 条结果)")
            else:
                _warn("未配置 / 无结果")
        except Exception:
            _warn("跳过 (未配置 API)")

    # ── 5. 合并统计 ──
    print(f"  {DIM}{'─' * 50}{R}")
    print(f"  {DIM}▸{R} {BOLD}待测主机{R} {G}{len(all_subs)}{R} 个")

    # ── 6. 并发探活 ──
    _step("并发探活")
    all_rows = []
    with ThreadPoolExecutor(max_workers=50) as pool:
        futures = {pool.submit(probe_host, host, 5.0, {200, 302, 403, 500, 502, 503}, True): host for host in all_subs}
        for future in as_completed(futures):
            try:
                all_rows.extend(future.result())
            except Exception:
                pass

    alive_hosts = len({row["host"] for row in all_rows if row.get("alive")})
    _ok(f"{alive_hosts} 存活 / {len(all_subs)} 总计")

    # ── 7. 端口扫描 ──
    port_scan_results = {}
    if USE_PORT_SCAN:
        custom_ports = None
        if args.ports:
            custom_ports = _parse_ports(args.ports)

        # 决定扫描目标列表
        if args.cidr_scan:
            # CIDR 模式：展开网段，跳过子域名探活，直接扫描 IP
            _step(f"CIDR 展开 ({args.cidr_scan})")
            cidr_ips = _expand_cidr(args.cidr_scan)
            _ok(f"{len(cidr_ips)} 个 IP")
            scan_targets = cidr_ips
        else:
            # 域名模式：只扫已探活的主机
            scan_targets = sorted({row["host"] for row in all_rows if row.get("alive")})

        if scan_targets:
            _step(f"端口扫描 ({len(scan_targets)} 主机)")
            try:
                scan_output = port_scan_hosts(scan_targets, ports=custom_ports)
                for host, (h, open_ports, svc) in scan_output.items():
                    if open_ports:
                        port_scan_results[host] = open_ports
            except Exception:
                # 降级到逐主机扫描
                _warn("批量扫描失败，降级为逐主机扫描")
                with ThreadPoolExecutor(max_workers=20) as pool:
                    futures = {pool.submit(port_scan_host, h, "tcp", custom_ports): h for h in scan_targets}
                    for future in as_completed(futures):
                        try:
                            host, ports, *_ = future.result()
                            if ports:
                                port_scan_results[host] = ports
                        except Exception:
                            pass
            if port_scan_results:
                total_ports = sum(len(p) for p in port_scan_results.values())
                _ok(f"{len(port_scan_results)} 主机开放 / {total_ports} 端口")
            else:
                _warn("无开放端口")

    # ── 8. IP 段查询 ──
    _step("IP 段查询")
    public_ips = _public_ipv4s(base_domain)
    ip138_ips = get_ips_from_ip138(f"www.{base_domain}")
    if not ip138_ips:
        ip138_ips = public_ips

    cidr = ""
    prefixes = []
    asn = None
    if ip138_ips:
        try:
            cidr = get_cidr_from_ip138(ip138_ips[0])
        except Exception:
            cidr = ""

    if USE_ASN:
        try:
            prefixes, asn = get_prefixes_for_domain(base_domain)
        except Exception:
            prefixes, asn = [], None

    if prefixes:
        asn_label = f"AS{asn} " if asn else ""
        _ok(f"{Y}{asn_label}{cidr or prefixes[0]}{G} ({len(prefixes)} 段){R}")
    elif asn:
        _ok(f"AS{asn} {cidr}" if cidr else f"AS{asn}")
    elif cidr:
        _ok(cidr)
    else:
        fallback_ip = ip138_ips[0] if ip138_ips else (public_ips[0] if public_ips else "")
        if fallback_ip:
            _warn(f"{fallback_ip} (未取到 CIDR)")
        else:
            _fail("未获取到")

    # ── 9. 反向 DNS ──
    scan_cidr = cidr or (prefixes[0] if prefixes else "")
    alive_ips = []
    ip_domain_rows = []
    if scan_cidr:
        _step("反向 DNS (rdnsdb)")
        try:
            ip_domain_rows = get_ip_domain_rows_from_rdnsdb(scan_cidr)
            _ok(f"{len(ip_domain_rows)} 条映射")
        except Exception:
            _fail("查询失败")

    # 资产 sheet
    fields = ["域名/IP", "IP", "URL", "HTTP状态码", "HTTP标题", "HTTPS状态码", "HTTPS标题"]
    grouped = {}
    grouped_urls = {}
    for row in all_rows:
        if not row.get("alive"):
            continue
        host = row.get("host", "")
        scheme = row.get("scheme")
        if scheme in ("http", "https") and host:
            grouped_urls.setdefault(host, set()).add(f"{scheme}://{host}")
        item = grouped.setdefault(
            host,
            {
                "域名/IP": host,
                "IP": ",".join(row.get("ips") or []),
                "URL": "",
                "HTTP状态码": "",
                "HTTP标题": "",
                "HTTPS状态码": "",
                "HTTPS标题": "",
            },
        )
        if scheme in ("http", "https"):
            prefix = scheme.upper()
            item[f"{prefix}状态码"] = row.get("status", "")
            item[f"{prefix}标题"] = row.get("title", "")

    for host, urls_set in grouped_urls.items():
        if host in grouped:
            grouped[host]["URL"] = _preferred_url_for_host(host, urls_set)

    # URL 备份 txt
    url_whitelist = {200, 301, 302, 303, 307, 308, 403}
    urls = set()
    for row in all_rows:
        if not row.get("alive"):
            continue
        try:
            status = int(row.get("status") or 0)
        except Exception:
            status = 0
        host = row.get("host")
        scheme = row.get("scheme")
        if scheme in ("http", "https") and host and status in url_whitelist:
            candidates = grouped_urls.get(host, set())
            preferred = _preferred_url_for_host(host, candidates)
            if preferred:
                urls.add(preferred)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "assets"
    ws1.append(fields)
    for host in sorted(grouped):
        row = grouped[host]
        ws1.append([row.get(field, "") for field in fields])

    ws2 = wb.create_sheet("ip_domains")
    ws2.append(["IP", "域名", "来源"])
    for ip, domain_name in ip_domain_rows:
        ws2.append([ip, domain_name, "rdnsdb"])

    ws3 = wb.create_sheet("asn_info")
    ws3.append(["项目", "值"])
    ws3.append(["ASN", asn if asn else ""])
    ws3.append(["CIDR (ip138)", cidr if cidr else ""])
    ws3.append(["解析IP", ",".join(ip138_ips) if ip138_ips else ""])
    if prefixes:
        ws3.append(["网段数", len(prefixes)])
        ws3.append([])
        ws3.append(["网段"])
        for p in prefixes:
            ws3.append([p])

    # FOFA 资产 sheet
    if fofa_results:
        ws4 = wb.create_sheet("fofa")
        fofa_fields = ["host", "ip", "domain", "port", "protocol", "title", "country", "city"]
        ws4.append(fofa_fields)
        for r in fofa_results:
            ws4.append([r.get(f, "") for f in fofa_fields])

    # 端口扫描 sheet
    if port_scan_results:
        ws5 = wb.create_sheet("port_scan")
        ws5.append(["主机", "开放端口"])
        for host in sorted(port_scan_results):
            ws5.append([host, ",".join(str(p) for p in port_scan_results[host])])

    try:
        wb.save(xlsx_file)
    except PermissionError:
        fallback_dir = os.path.join(os.path.expanduser("~"), f"{base_domain}_output")
        os.makedirs(fallback_dir, exist_ok=True)
        xlsx_file = os.path.join(fallback_dir, f"{base_domain}_assets.xlsx")
        wb.save(xlsx_file)
        output_dir = fallback_dir
    except Exception as e:
        print(f"\n  {RED}Failed to write Excel: {e}{R}\n")
        sys.exit(1)

    try:
        urls_file = os.path.join(output_dir, f"{base_domain}_urls.txt")
        with open(urls_file, "w", encoding="utf-8") as uf:
            for url in sorted(urls):
                uf.write(url + "\n")
    except Exception:
        pass

    try:
        ip_domains_file = os.path.join(output_dir, f"{base_domain}_ip_domains.txt")
        with open(ip_domains_file, "w", encoding="utf-8") as df:
            for ip, domain_name in ip_domain_rows:
                df.write(f"{ip}\t{domain_name}\n")
    except Exception:
        pass

    print(f"\n  {DIM}{'─' * 50}{R}")
    print(f"  {BOLD}📂 输出文件{R}")
    print(f"    {Y}Excel{R}  {os.path.abspath(xlsx_file)}")
    if urls:
        print(f"    {Y}URLs{R}   {os.path.abspath(urls_file)}")
    print(f"    {Y}目录{R}    {os.path.abspath(output_dir)}")
    print(f"  {DIM}{'─' * 50}{R}\n")


if __name__ == "__main__":
    main()